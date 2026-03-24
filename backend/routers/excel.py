import os
import tempfile
from datetime import datetime

from fastapi import APIRouter, Depends, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from database import get_db
import models

from openpyxl import Workbook

router = APIRouter(prefix="/excel", tags=["Excel"])


# ==========================================
# EXPORTAR TODO EL SISTEMA A EXCEL
# ==========================================

def _cleanup_file(path: str):
    try:
        os.remove(path)
    except OSError:
        pass


@router.get("/exportar")
def exportar_excel(background_tasks: BackgroundTasks, db: Session = Depends(get_db)):

    wb = Workbook()

    # =================================================
    # HOJA 1 → ASIGNACIONES (UNA FILA POR SERVICIO/FECHA)
    # =================================================

    ws = wb.active
    ws.title = "Asignaciones por Servicio"

    ws.append([
        "Fecha", "Servicio", "Complejo", "Temática", "Empresa", "Grupo", 
        "Check-in", "Check-out", 
        "Estudiantes", "Padres", "Guias", "Total Pax",
        "Alcohol Grupo", "Con Comida"
    ])

    # Obtener todas las asignaciones y ordenarlas por fecha
    asignaciones = db.query(models.Asignacion).all()
    asignaciones_sorted = sorted(
        asignaciones,
        key=lambda x: (
            x.fecha_evento.fecha if x.fecha_evento else datetime.max.date(),
            x.fecha_evento.evento.nombre if x.fecha_evento and x.fecha_evento.evento else ""
        )
    )

    for a in asignaciones_sorted:
        grupo = a.grupo
        fecha_evento = a.fecha_evento
        evento = fecha_evento.evento if fecha_evento else None

        if not grupo or not fecha_evento or not evento:
            continue
        
        # Determinar si tiene comida según el tipo de servicio
        con_comida = ""
        if evento.tipo == "PARQUE":
            con_comida = "SI" if grupo.parque_con_comida else "NO"
        elif evento.tipo == "POOL":
            con_comida = "SI" if grupo.pool_con_comida else "NO"
        else:  # DISCO
            con_comida = "N/A"
        
        # Temática (solo para discos)
        tematica = fecha_evento.tematica.nombre if fecha_evento.tematica else ""
        
        ws.append([
            fecha_evento.fecha,
            evento.nombre,
            evento.complejo or "",
            tematica,
            grupo.empresa.nombre,
            grupo.nombre,
            grupo.fecha_entrada,
            grupo.fecha_salida,
            grupo.cantidad_estudiantes,
            grupo.cantidad_padres,
            grupo.cantidad_guias,
            grupo.cantidad_pax,
            "SI" if grupo.permite_alcohol else "NO",
            con_comida
        ])

    # =================================================
    # HOJA 2 → DISCOS
    # =================================================

    ws2 = wb.create_sheet("Discos")

    ws2.append(["Fecha", "Disco", "Complejo", "Temática", "Alcohol", "Pax total"])

    fechas = db.query(models.FechaEvento).all()

    for f in fechas:
        if f.evento.tipo != "DISCO":
            continue

        total = sum(a.grupo.cantidad_pax for a in f.asignaciones if a.grupo)
        tematica = f.tematica.nombre if f.tematica else ""

        ws2.append([
            f.fecha,
            f.evento.nombre,
            f.evento.complejo or "",
            tematica,
            "CON" if f.con_alcohol else "SIN",
            total
        ])

    # =================================================
    # HOJA 3 → PARQUE (POR EMPRESA)
    # =================================================

    ws3 = wb.create_sheet("Parque")

    ws3.append(["Fecha", "Empresa", "Pax con comida", "Pax sin comida"])

    for f in fechas:
        if f.evento.tipo != "PARQUE":
            continue

        empresas = {}

        for a in f.asignaciones:
            if not a.grupo or not a.grupo.empresa:
                continue
            empresa = a.grupo.empresa.nombre

            if empresa not in empresas:
                empresas[empresa] = {"con": 0, "sin": 0}

            if a.grupo.parque_con_comida:
                empresas[empresa]["con"] += a.grupo.cantidad_pax
            else:
                empresas[empresa]["sin"] += a.grupo.cantidad_pax

        for emp, datos in empresas.items():
            ws3.append([f.fecha, emp, datos["con"], datos["sin"]])

    # =================================================
    # HOJA 4 → POOL (POR EMPRESA)
    # =================================================

    ws4 = wb.create_sheet("Pool")

    ws4.append(["Fecha", "Empresa", "Pax con comida", "Pax sin comida"])

    for f in fechas:
        if f.evento.tipo != "POOL":
            continue

        empresas = {}

        for a in f.asignaciones:
            if not a.grupo or not a.grupo.empresa:
                continue
            empresa = a.grupo.empresa.nombre

            if empresa not in empresas:
                empresas[empresa] = {"con": 0, "sin": 0}

            if a.grupo.pool_con_comida:
                empresas[empresa]["con"] += a.grupo.cantidad_pax
            else:
                empresas[empresa]["sin"] += a.grupo.cantidad_pax

        for emp, datos in empresas.items():
            ws4.append([f.fecha, emp, datos["con"], datos["sin"]])

    # =================================================
    # GUARDAR
    # =================================================

    filename = f"reporte_eventos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        filepath = tmp.name

    wb.save(filepath)
    background_tasks.add_task(_cleanup_file, filepath)

    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==========================================
# EXPORTAR REPORTE FINANCIERO DE UNA EMPRESA
# ==========================================

# ==========================================
# EXPORTAR RESUMEN FINANCIERO DE TODAS
# ==========================================

@router.get("/finanzas/todas")
def exportar_finanzas_todas(background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    from routers.finanzas import get_resumen_empresa

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Empresas"

    ws.append([
        "Empresa",
        "Total PAX",
        "Total Cobrado",
        "Saldo",
        "Total a Pagar"
    ])

    empresas = db.query(models.Empresa).all()
    for e in empresas:
        res = get_resumen_empresa(e.id, db)
        total_pax = sum(g.get("pax", 0) for g in res.get("grupos", []))
        ws.append([
            e.nombre,
            total_pax,
            res["total_pagado"],
            res["saldo"],
            res["total_venta"]
        ])

    filename = f"reporte_financiero_todas_{datetime.now().strftime('%Y%m%d')}.xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        filepath = tmp.name

    wb.save(filepath)
    background_tasks.add_task(_cleanup_file, filepath)

    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get("/finanzas/{empresa_id:int}")
def exportar_finanzas(empresa_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    from routers.finanzas import get_resumen_empresa
    
    res = get_resumen_empresa(empresa_id, db)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Financiero"
    
    ws.append(["REPORTE FINANCIERO - " + res["empresa"]])
    ws.append([])
    
    # Detalle de grupos
    ws.append(["Grupo", "Servicio", "Detalle", "Precio Unit.", "Cant.", "Estud.", "Padres", "Guias", "Total PAX", "Pagantes", "Subtotal"])
    for g in res["grupos"]:
        # Fila de encabezado de grupo
        ws.append([g["nombre"], "", "", "", "", g["estudiantes"], g["padres"], g["guias"], g["pax"], "", g["subtotal"]])
        # Filas de detalles de servicios
        for s in g["servicios"]:
            ws.append([
                "", 
                s["servicio"], 
                s["descripcion"], 
                s["precio_u"], 
                s["cantidad"], 
                s["estudiantes"],
                s["padres"],
                s["guias"],
                s["pax_original"],
                s["pax"], # Este es el 'pax_cobrar'
                s["subtotal"]
            ])
    
    ws.append([])
    ws.append(["TOTAL VENTA", res["total_venta"]])
    ws.append(["TOTAL COBRADO", res["total_pagado"]])
    ws.append(["SALDO", res["saldo"]])
    
    ws.append([])
    ws.append(["HISTORIAL DE PAGOS"])
    ws.append(["Fecha", "Monto", "Método", "Nota"])
    
    pagos = db.query(models.Pago).filter(models.Pago.empresa_id == empresa_id).all()
    for p in pagos:
        ws.append([p.fecha, p.monto, p.metodo, p.nota])
        
    filename = f"reporte_financiero_{res['empresa']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        filepath = tmp.name

    wb.save(filepath)
    background_tasks.add_task(_cleanup_file, filepath)
    
    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
